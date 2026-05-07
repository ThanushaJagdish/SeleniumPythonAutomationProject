import json
import csv
from openpyxl import load_workbook


# Excel Data Provider
def read_excel_data(filepath, sheetname):
    wb = load_workbook(filepath)  # Load the Excel workbook.
    sheet = wb[sheetname]  # Access the specified sheet by name.
    data = []  # Empty list to store data rows
    for row in sheet.iter_rows(min_row=2,
                               values_only=True):  # Loop through all rows from the 2nd row onwards (to skip the header)
        data.append(row)  # Add each row (as a tuple) to the list
    return data  # Return list of tuples